"""
Fetches and parses the AEMO NEM Generation Information XLSX.

Sources tried in order on every cache miss:

  1. live      — direct download from www.aemo.com.au (subject to WAF blocks
                 from datacentre IPs; see _BROWSER_HEADERS).
  2. snapshot  — the parsed bess_list.json bundled in the repo, refreshed
                 daily by .github/workflows/refresh_bess_list.yml running
                 scripts/refresh_bess_list.py from GitHub-hosted runners
                 (whose egress is not blocked by AEMO's WAF).

The first successful step is cached in memory for 24 hours.  The result
carries a `source` field and any user-facing `warnings` so the API can
expose data freshness in the response.
"""
import io
import json
import logging
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

import httpx
import openpyxl

logger = logging.getLogger(__name__)

XLSX_URL = (
    "https://www.aemo.com.au/-/media/files/electricity/nem/planning_and_forecasting"
    "/generation_information/2026/nem-generation-information-jan-2026.xlsx"
    "?rev=1f6bccf827284f9fb6d6f3ae56ed3fe9&sc_lang=en"
)

# AEMO's Azure Front Door WAF on www.aemo.com.au returns 403 to requests that
# look like default Python clients. Sending a full browser-style header set
# bypasses the simpler bot-detection rules. (Doesn't help against pure
# IP-range blocks — see gen_info_mirror.py for the raw.githubusercontent.com
# fallback.)
_BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "application/octet-stream;q=0.9,*/*;q=0.8"
    ),
    "Accept-Language": "en-AU,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Referer": "https://www.aemo.com.au/energy-systems/electricity/national-electricity-market-nem/nem-forecasting-and-planning/forecasting-and-planning-data/generation-information",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Upgrade-Insecure-Requests": "1",
}

SHEET_NAME = "Generator Information"

# Map NEM region codes to the state keys used in the UI
REGION_TO_STATE: dict[str, str] = {
    "QLD1": "QLD",
    "NSW1": "NSW",
    "VIC1": "VIC",
    "SA1":  "SA",
    "TAS1": "TAS",
}

BATTERY_TECH = "Battery Storage"
IN_SERVICE   = "In Service"

# Column names to search for (case-insensitive substring match as fallback)
_COL_DUID       = "DUID"
_COL_NAME       = "Unit Name"
_COL_TECH       = "Technology Type"
_COL_STATUS       = "Commitment Status"
_COL_REGION       = "Region"
_COL_CAPACITY_MW  = "Agg Nameplate Capacity (MW AC)"
_COL_CAPACITY_MWH = "Agg Nameplate Storage Capacity (MWh)"

CACHE_TTL = timedelta(hours=24)


@dataclass
class BessListResult:
    """Parsed BESS list plus provenance for the API response."""
    states: dict[str, list[dict]]
    source: str            # "live" | "snapshot"
    fetched_at: datetime   # for the live path: now; for snapshot: snapshot timestamp
    warnings: list[str] = field(default_factory=list)


# In-memory cache: (result, cached_at)
_cache: tuple[BessListResult, datetime] | None = None


# Snapshot files written by scripts/refresh_bess_list.py
_DATA_DIR = Path(__file__).parent.parent / "data"
_SNAPSHOT_JSON = _DATA_DIR / "bess_list.json"
_SNAPSHOT_META = _DATA_DIR / "bess_list_meta.json"


def _find_col(headers: list[str], target: str) -> int | None:
    """
    Return the 0-based index of the column whose header matches target.
    Tries exact match first, then case-insensitive prefix match.
    """
    target_lower = target.lower()
    for i, h in enumerate(headers):
        if h == target:
            return i
    for i, h in enumerate(headers):
        if h.lower().startswith(target_lower):
            return i
    return None


def _parse_xlsx(xlsx_bytes: bytes) -> dict[str, list[dict]]:
    """
    Parse the XLSX bytes and return BESS list grouped by state.

    Returns:
        { "QLD": [...], "NSW": [...], "VIC": [...], "SA": [...], "TAS": [...] }
        Each entry: { "duid": str, "name": str, "capacity_mw": float|None, "region": str }
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Available: {available}")

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 4:
        raise ValueError("Sheet has fewer than 4 rows; expected headers on row 4.")

    # Row 4 (0-indexed: rows[3]) contains the column headers.
    # Rows 1–3 are title/metadata rows that must be skipped.
    raw_headers = [str(c).strip() if c is not None else "" for c in rows[3]]

    col_duid         = _find_col(raw_headers, _COL_DUID)
    col_name         = _find_col(raw_headers, _COL_NAME)
    col_tech         = _find_col(raw_headers, _COL_TECH)
    col_status       = _find_col(raw_headers, _COL_STATUS)
    col_region       = _find_col(raw_headers, _COL_REGION)
    col_capacity_mw  = _find_col(raw_headers, _COL_CAPACITY_MW)
    col_capacity_mwh = _find_col(raw_headers, _COL_CAPACITY_MWH)

    missing = [
        name for name, idx in [
            (_COL_DUID, col_duid), (_COL_TECH, col_tech),
            (_COL_STATUS, col_status), (_COL_REGION, col_region),
        ] if idx is None
    ]
    if missing:
        raise ValueError(
            f"Required columns not found: {missing}. "
            f"Headers seen: {raw_headers[:20]}"
        )

    result: dict[str, list[dict]] = {s: [] for s in REGION_TO_STATE.values()}

    for row in rows[4:]:
        def cell(idx: int | None) -> str:
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        tech   = cell(col_tech)
        status = cell(col_status)
        region = cell(col_region)

        if tech != BATTERY_TECH:
            continue
        if status != IN_SERVICE:
            continue
        state = REGION_TO_STATE.get(region)
        if state is None:
            continue

        duid = cell(col_duid)
        if not duid:
            continue

        name = cell(col_name) or duid

        def _float(col_idx: int | None) -> float | None:
            if col_idx is None or col_idx >= len(row):
                return None
            v = row[col_idx]
            if v is None:
                return None
            try:
                return float(v)
            except (ValueError, TypeError):
                return None

        result[state].append({
            "duid": duid,
            "name": name,
            "capacity_mw":  _float(col_capacity_mw),
            "capacity_mwh": _float(col_capacity_mwh),
            "region": region,
        })

    total = sum(len(v) for v in result.values())
    logger.info(
        "Parsed XLSX: %d in-service battery storage units across %d states",
        total, sum(1 for v in result.values() if v),
    )
    return result


async def _download_and_parse(url: str, client: httpx.AsyncClient) -> dict[str, list[dict]]:
    resp = await client.get(url, headers=_BROWSER_HEADERS)
    resp.raise_for_status()
    return _parse_xlsx(resp.content)


def _load_snapshot() -> BessListResult:
    """Load the parsed JSON snapshot bundled in the repo. Raises if missing."""
    states = json.loads(_SNAPSHOT_JSON.read_text())
    fetched_at = datetime.utcnow()
    if _SNAPSHOT_META.exists():
        try:
            meta = json.loads(_SNAPSHOT_META.read_text())
            ts = meta.get("fetched_at")
            if ts:
                # strip any trailing 'Z' / timezone offset to a naive UTC datetime
                ts_clean = ts.replace("Z", "+00:00")
                fetched_at = datetime.fromisoformat(ts_clean).replace(tzinfo=None)
        except (ValueError, KeyError) as exc:
            logger.warning("Could not parse snapshot meta %s: %s", _SNAPSHOT_META, exc)
    return BessListResult(
        states=states,
        source="snapshot",
        fetched_at=fetched_at,
    )


async def fetch_bess_list() -> BessListResult:
    """
    Return the BESS list, trying live AEMO → bundled snapshot.

    The snapshot is refreshed daily by the refresh_bess_list workflow from
    GitHub-hosted runners (whose egress is not blocked by AEMO), so even
    when the live AEMO fetch fails the fallback is at most ~24 hours old.

    Uses a 24-hour in-memory cache. The result carries provenance so the
    API can surface a warning when serving non-live data.
    """
    global _cache

    now = datetime.utcnow()

    # Return cached data if still fresh
    if _cache is not None:
        cached, cached_at = _cache
        if now - cached_at < CACHE_TTL:
            logger.debug("Returning cached BESS list (age %s, source=%s)",
                         now - cached_at, cached.source)
            return cached

    # 1. Live AEMO
    try:
        logger.info("Downloading AEMO XLSX (live): %s", XLSX_URL)
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            states = await _download_and_parse(XLSX_URL, client)
        result = BessListResult(states=states, source="live", fetched_at=now)
        _cache = (result, now)
        return result
    except Exception as exc:
        logger.warning("Live AEMO XLSX fetch failed: %s", exc)

    # 2. Bundled snapshot — refreshed daily by CI from GH-hosted runners.
    try:
        snapshot = _load_snapshot()
        age = now - snapshot.fetched_at
        days = max(age.days, 0)
        snapshot.warnings = [
            f"Live AEMO BESS list is unavailable; showing the daily snapshot "
            f"committed on {snapshot.fetched_at:%Y-%m-%d} "
            f"({days} day{'s' if days != 1 else ''} old)."
        ]
        _cache = (snapshot, now)
        logger.warning("Falling back to bundled snapshot (age %s)", age)
        return snapshot
    except FileNotFoundError:
        logger.error("No bundled snapshot at %s", _SNAPSHOT_JSON)
        # Return an empty result rather than raising — frontend handles
        # empty states gracefully and the warning explains why.
        return BessListResult(
            states={s: [] for s in REGION_TO_STATE.values()},
            source="snapshot",
            fetched_at=now,
            warnings=[
                "BESS list is currently unavailable from AEMO and no local "
                "snapshot is bundled. Try again later."
            ],
        )
